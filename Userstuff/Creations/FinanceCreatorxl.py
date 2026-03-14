import openpyxl as xl

'''expense tracker'''
print("_expense tracker_")
Total_amount = int(input("Cash in hand: "))
Date = input("Enter today's date: ")
Name = input("What is the name of the file you want to create: ")
filename = f"{Name}.xlsx"  

'''expenses'''
print("How much money have you spent on:")
Grocery = int(input("Grocery: "))
vegetables = int(input("Vegetables: "))
milk_products = int(input("Milk Products: "))
stationary = int(input("Stationary: "))
snacks = int(input("Snacks: "))
zomato = int(input("Zomato: "))
card = input("Which card used? [cr/dr]: ")

'''Calc'''
final_val = Grocery + vegetables + milk_products + stationary + snacks + zomato
sub_amount = Total_amount - final_val
if sub_amount < 0:
    print(f"Max budget {Total_amount} reached")
elif sub_amount == 0:
    print(f"Budget finished")
else:
    print(f"Budget: {Total_amount}, Amount spent: {final_val}, Amount left: {sub_amount}")

wb = xl.Workbook()
sheet = wb.active
sheet.title = "Finance"

# headers
sheet["A1"] = "Date"
sheet["B1"] = "Description"
sheet["C1"] = "Amount"
sheet["D1"] = "Cr/dr"


items = ["Cash in Hand","Grocery", "Vegetables", "Milk products", "Stationary", "Snacks", "Zomato"]
values = [Total_amount ,Grocery, vegetables, milk_products, stationary, snacks, zomato]

current_row = 2
for item, val in zip(items, values):
    sheet.cell(row=current_row, column=1, value=Date)    
    sheet.cell(row=current_row, column=2, value=item)     
    sheet.cell(row=current_row, column=3, value=val)        
    sheet.cell(row=current_row, column=4, value=card)       
    current_row += 1

sheet.cell(row=current_row, column=2, value="Total spent")  
sheet.cell(row=current_row, column=3, value=final_val) 

current_row += 1

sheet.cell(row=current_row, column=2, value="Balance")
sheet.cell(row=current_row, column=3, value=sub_amount)

wb.save(filename)
print(f"{filename} created ")

